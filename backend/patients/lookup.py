from __future__ import annotations

import json
import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class PatientLookupService:
    """Join en memoria entre los dos datasets de perfil de paciente que ya trae
    el reto (`dataset/perfiles_pacientes_co.xlsx` para identidad y
    `dataset/perfiles_clinicos_pacientes_silver_contest.xlsx` para el
    procedimiento), unidos por `paciente_id`.

    Antes del punto 1.2 del plan, nadie usaba este dato en el flujo de llamada:
    el agente no sabía a quién llamaba ni de qué cirugía. Es exactamente el
    hueco que señaló el jurado en 3.2 ("no hay identidad de paciente ni de
    procedimiento en ninguna parte del flujo de llamada").

    Vive como catálogo de solo lectura (se carga una vez, no cambia por
    sesión) — igual que CorpusVectorStore, se cachea como singleton en
    main.py para no releer los .xlsx en cada request.
    """

    def __init__(self, root_dir: Path) -> None:
        self.root_dir = root_dir
        self.dataset_dir = root_dir / "dataset"
        self._demographics: list[dict[str, object]] | None = None
        self._clinical: list[dict[str, object]] | None = None

    @staticmethod
    def _read_rows(path: Path) -> list[dict[str, object]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            headers = [str(header) for header in next(rows)]
            return [dict(zip(headers, values)) for values in rows]
        finally:
            workbook.close()

    def _load(self) -> None:
        if self._demographics is not None and self._clinical is not None:
            return
        self._demographics = self._read_rows(self.dataset_dir / "perfiles_pacientes_co.xlsx")
        self._clinical = self._read_rows(self.dataset_dir / "perfiles_clinicos_pacientes_silver_contest.xlsx")

    def sample_patients(self, limit: int = 5) -> list[dict[str, object]]:
        """Unos cuantos pacientes de ejemplo (id + nombre + procedimiento) para
        poblar el selector del frontend en la demo — no expone el dataset
        completo, solo lo mínimo necesario para elegir a quién "llamar"."""
        self._load()
        demographics_by_id = {
            str(row.get("paciente_id")): row for row in self._demographics or []
        }
        sample = []
        for clinical in self._clinical or []:
            demographic = demographics_by_id.get(str(clinical.get("paciente_id")))
            if demographic is None:
                continue
            sample.append((clinical, demographic))
            if len(sample) >= limit:
                break
        return [
            {
                "paciente_id": str(clinical["paciente_id"]),
                "nombre_completo": str(demographic["nombre_completo"]),
                "procedimiento": str(clinical["procedimiento"]),
            }
            for clinical, demographic in sample
        ]

    def get_patient_context(self, paciente_id: str) -> dict[str, object] | None:
        """None si el paciente_id no existe en alguno de los dos datasets —
        el llamador decide si eso es un 404 o simplemente seguir sin contexto
        de paciente (comportamiento aditivo: la llamada sigue funcionando sin
        identidad si no se encuentra)."""
        self._load()
        demo = next(
            (row for row in self._demographics or [] if str(row.get("paciente_id")) == paciente_id),
            None,
        )
        clinical = next(
            (row for row in self._clinical or [] if str(row.get("paciente_id")) == paciente_id),
            None,
        )
        if demo is None or clinical is None:
            logger.warning("paciente_id '%s' no encontrado en ambos datasets, sigue sin contexto", paciente_id)
            return None

        try:
            comorbilidades = json.loads(str(clinical.get("comorbilidades") or "[]"))
        except (TypeError, ValueError):
            comorbilidades = []

        fecha_cirugia_raw = clinical.get("fecha_cirugia")
        if isinstance(fecha_cirugia_raw, (datetime, date)):
            fecha_cirugia = fecha_cirugia_raw.isoformat()[:10]
        else:
            fecha_cirugia = str(fecha_cirugia_raw)[:10] if fecha_cirugia_raw is not None else ""

        edad_raw = clinical.get("edad")
        edad = int(edad_raw) if edad_raw is not None else None

        # Días transcurridos desde la cirugía. Es el dato que permite al motor
        # de decisión distinguir un signo leve esperable en la recuperación
        # temprana de uno que a la semana ya no lo es. En el gold-set el día
        # viene explícito (columna `dia_postop`); en una llamada real hay que
        # calcularlo, y esto es lo que lo calcula.
        dias_postop = None
        if fecha_cirugia:
            try:
                dias_postop = (date.today() - date.fromisoformat(fecha_cirugia)).days
            except ValueError:
                dias_postop = None

        return {
            "paciente_id": paciente_id,
            "nombre_completo": str(demo.get("nombre_completo", "")),
            "dias_postop": dias_postop,
            "procedimiento": str(clinical.get("procedimiento", "")),
            "fecha_cirugia": fecha_cirugia,
            "edad": edad,
            "genero": str(clinical.get("genero") or "") or None,
            "comorbilidades": comorbilidades,
            # Valor crudo de modulo_synthea (ej. "appendicitis"). Se pasa al
            # prompt como pista de categoría clínica, pero NO se usa para
            # filtrar duro la búsqueda RAG: los nombres de carpeta del corpus
            # (p.ej. "Appendicitis", "colorectal cancer" con espacio) no
            # coinciden 1:1 en formato con este valor, y un filtro mal
            # emparejado dejaría resultados vacíos en vez de degradar
            # gradualmente. Queda como mejora futura documentada, no como
            # riesgo asumido a días de la demo en vivo.
            "categoria_clinica": str(clinical.get("modulo_synthea", "")),
        }
